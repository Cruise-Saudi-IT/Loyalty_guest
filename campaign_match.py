"""Match May26 Repeaters segment CSVs to reservations created/changed after
2026-05-14 17:00 KSA (= 14:00 system time, system is UTC = KSA - 3h).

Uses SEAWARE.RES_EVENT to detect creation events and OF→BK transfers,
joined to GUEST_INFO for phone match, and SEAWARE.RES_PARAM to check
whether the campaign referral code was attached to the reservation.

Referral attribution lives in SEAWARE.RES_PARAM:
  PARAM_CODE  = 'REFERRAL_SOURCE'  (or REFERRAL_TYPE — same value)
  PARAM_VALUE = e.g. 'ARYSAIL526'
  RES_ID      = the reservation it was applied to

Referral codes per segment (per the email):
  Repeat-Repeater + Premium Repeater  → ARYPREM526
  Family Repeater                     → ARYFAM526
  RedSea Loyalist + Recent Repeater   → ARYCLUB526
  Repeater                            → ARYSAIL526
"""
import os
import csv
import datetime
from collections import defaultdict
import oracledb
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CSV_DIR = r'C:\Users\RafalASobahi\OneDrive - Cruise Saudi'
OUT_DIR = CSV_DIR
CUTOFF = '2026-05-14 14:00:00'  # system time; KSA = +3h = 17:00

SEGMENT_FILES = [
    ('May26 Repeat-Repeater Segments.csv', 'Repeat-Repeater'),
    ('May26 Premium Repeater Segments.csv', 'Premium Repeater'),
    ('May26 Family Repeater Segments.csv', 'Family Repeater'),
    ('May26 RedSea Loyalist Segments.csv', 'RedSea Loyalist'),
    ('May26 Recent Repeater Segments.csv', 'Recent Repeater'),
    ('May26 Repeater Segments.csv', 'Repeater'),
]

SEGMENT_CODES = {
    'Repeat-Repeater':  'ARYPREM526',
    'Premium Repeater': 'ARYPREM526',
    'Family Repeater':  'ARYFAM526',
    'RedSea Loyalist':  'ARYCLUB526',
    'Recent Repeater':  'ARYCLUB526',
    'Repeater':         'ARYSAIL526',
}

# Reverse map: code -> list of segments it represents (per the email's table)
CODE_SEGMENTS = {
    'ARYPREM526': ['Repeat-Repeater', 'Premium Repeater'],
    'ARYFAM526':  ['Family Repeater'],
    'ARYCLUB526': ['RedSea Loyalist', 'Recent Repeater'],
    'ARYSAIL526': ['Repeater'],
}
ALL_CAMPAIGN_CODES = list(CODE_SEGMENTS.keys())


def normalize_phone(raw):
    digits = ''.join(c for c in raw if c.isdigit())
    return digits[-9:] if len(digits) >= 9 else ''


def load_segments():
    phone_to_segments = defaultdict(set)
    seg_counts = {}
    for fname, label in SEGMENT_FILES:
        path = os.path.join(CSV_DIR, fname)
        phones = set()
        with open(path, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            next(reader, None)
            for row in reader:
                if not row:
                    continue
                p9 = normalize_phone(row[0])
                if p9:
                    phones.add(p9)
                    phone_to_segments[p9].add(label)
        seg_counts[label] = len(phones)
    return phone_to_segments, seg_counts


def fetch_events_and_referrals(all_phones):
    """Return (events, referrals_by_resid).
    events  : list of (res_id, event_type, old_status, new_status,
                       event_ts, phone, name, init_date, pkg, sail,
                       cur_status)
    referrals_by_resid : dict res_id -> set of referral codes from RES_PARAM
    """
    conn = oracledb.connect(user='aryadmin', password='Cruisesaudi_2021',
                            dsn='localhost:15210/AROYAREP')
    cur = conn.cursor()

    # Events × matched guests
    events = []
    for i in range(0, len(all_phones), 900):
        batch = all_phones[i:i + 900]
        binds = {f'p{j}': v for j, v in enumerate(batch)}
        ph = ','.join(f':p{j}' for j in range(len(batch)))
        sql = f"""
            SELECT e.RES_ID,
                   e.RES_EVENT_TYPE,
                   e.OLD_STATUS,
                   e.NEW_STATUS,
                   e.EVENT_TIMESTAMP,
                   SUBSTR(TRIM(g.PHONE_NUMBER), -9) AS phone,
                   g.FULL_NAME,
                   g.RES_INIT_DATE,
                   g.PACKAGE_TYPE,
                   g.SAIL_DATE_FROM,
                   g.RES_STATUS,
                   g.EMAIL
            FROM SEAWARE.RES_EVENT e
            JOIN GUEST_INFO g ON e.RES_ID = g.RES_ID
            WHERE e.EVENT_TIMESTAMP > TO_DATE(:cutoff, 'YYYY-MM-DD HH24:MI:SS')
              AND SUBSTR(TRIM(g.PHONE_NUMBER), -9) IN ({ph})
        """
        cur.execute(sql, {**binds, 'cutoff': CUTOFF})
        events.extend(cur.fetchall())

    # Referral codes per RES_ID — RES_PARAM stores referral attribution
    # under PARAM_CODE='REFERRAL_SOURCE' (and a duplicate REFERRAL_TYPE row).
    res_ids = list({e[0] for e in events if e[0] is not None})
    referrals_by_resid = defaultdict(set)
    for i in range(0, len(res_ids), 900):
        batch = res_ids[i:i + 900]
        binds = {f'r{j}': v for j, v in enumerate(batch)}
        ph = ','.join(f':r{j}' for j in range(len(batch)))
        sql = f"""
            SELECT RES_ID, PARAM_VALUE
            FROM SEAWARE.RES_PARAM
            WHERE PARAM_CODE = 'REFERRAL_SOURCE'
              AND PARAM_VALUE IS NOT NULL
              AND RES_ID IN ({ph})
        """
        cur.execute(sql, binds)
        for res_id, code in cur.fetchall():
            referrals_by_resid[res_id].add(code.strip().upper())

    cur.close()
    conn.close()
    return events, referrals_by_resid


def fetch_all_code_bookings():
    """Every booking tagged with one of the 4 campaign codes,
    regardless of whether the guest's phone is in our segment CSVs.
    Returns list of dicts (one row per guest on the booking).
    """
    conn = oracledb.connect(user='aryadmin', password='Cruisesaudi_2021',
                            dsn='localhost:15210/AROYAREP')
    cur = conn.cursor()
    placeholders = ','.join(f':c{i}' for i in range(len(ALL_CAMPAIGN_CODES)))
    binds = {f'c{i}': v for i, v in enumerate(ALL_CAMPAIGN_CODES)}
    cur.execute(f"""
        SELECT DISTINCT
               rp.PARAM_VALUE       AS code,
               rp.RES_ID,
               g.FULL_NAME,
               SUBSTR(TRIM(g.PHONE_NUMBER), -9) AS phone,
               g.RES_STATUS,
               g.RES_INIT_DATE,
               g.PACKAGE_TYPE,
               g.SAIL_DATE_FROM,
               g.EMAIL
        FROM   SEAWARE.RES_PARAM rp
        JOIN   GUEST_INFO g ON g.RES_ID = rp.RES_ID
        WHERE  rp.PARAM_CODE = 'REFERRAL_SOURCE'
          AND  UPPER(rp.PARAM_VALUE) IN ({placeholders})
        ORDER  BY rp.PARAM_VALUE, rp.RES_ID, g.FULL_NAME
    """, binds)
    rows = []
    for r in cur.fetchall():
        rows.append({
            'code': r[0],
            'res_id': r[1],
            'name': r[2],
            'phone': r[3],
            'status': r[4],
            'init_date': r[5],
            'pkg': r[6],
            'sail': r[7],
            'email': r[8],
        })
    cur.close()
    conn.close()
    return rows


def main():
    phone_to_segments, seg_counts = load_segments()
    all_phones = list(phone_to_segments.keys())
    print(f'Loaded {len(all_phones)} unique phones across {len(SEGMENT_FILES)} segments')

    events, promos = fetch_events_and_referrals(all_phones)
    print(f'Fetched {len(events)} event-guest rows; {len(promos)} RES_IDs have a referral source')

    # Also fetch ALL bookings tagged with the 4 campaign codes regardless of
    # whether the guest's phone is in our segment CSVs — captures bookings
    # made via forwarded links by people outside the targeted lists.
    code_bookings = fetch_all_code_bookings()  # list of dicts

    # Per-RES_ID aggregation
    res_info = {}
    for (res_id, ev_type, old_s, new_s, ts, phone,
         name, init_date, pkg, sail, cur_status, email) in events:
        if res_id is None:
            continue
        rec = res_info.setdefault(res_id, {
            'phone': phone,
            'name': name,
            'email': email,
            'init_date': init_date,
            'pkg': pkg,
            'sail': sail,
            'cur_status': cur_status,  # from GUEST_INFO — authoritative current state
            'latest_event_ts': None,
            'latest_event_new_status': None,
            'created_in_window': False,
            'transferred_to_bk': False,
            'To_BK': None,         # when OF→BK transition happened
            'event_types': set(),
        })
        rec['event_types'].add(ev_type)
        if rec['latest_event_ts'] is None or (ts and ts > rec['latest_event_ts']):
            rec['latest_event_ts'] = ts
            rec['latest_event_new_status'] = new_s
        if ev_type == 'NEW BOOKING':
            rec['created_in_window'] = True
        if old_s == 'OF' and new_s == 'BK':
            rec['transferred_to_bk'] = True
            # Keep earliest OF→BK timestamp (the actual conversion moment)
            if rec['To_BK'] is None or (ts and ts < rec['To_BK']):
                rec['To_BK'] = ts

    # Bucket by segment
    seg_metrics = defaultdict(lambda: {
        'created': set(),
        'cur_bk': set(),
        'cur_of': set(),
        'cur_cx': set(),
        'transferred_to_bk': set(),
        'with_campaign_code': set(),
    })
    for res_id, rec in res_info.items():
        cur = rec['cur_status'] or rec['latest_event_new_status']
        for seg in phone_to_segments.get(rec['phone'], ()):
            m = seg_metrics[seg]
            if rec['created_in_window']:
                m['created'].add(res_id)
            if cur == 'BK':
                m['cur_bk'].add(res_id)
            elif cur == 'OF':
                m['cur_of'].add(res_id)
            elif cur == 'CX':
                m['cur_cx'].add(res_id)
            if rec['transferred_to_bk']:
                m['transferred_to_bk'].add(res_id)
            expected = SEGMENT_CODES.get(seg)
            applied = promos.get(res_id, set())
            if expected and expected in applied:
                m['with_campaign_code'].add(res_id)

    # ── Build single Excel workbook with three sheets ─────────────
    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.freeze_panes = 'A2'

    def autosize(ws):
        for col_cells in ws.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0
                           for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 2, 50)

    # Sheet 1 — Summary
    ws = wb.active
    ws.title = 'Summary'
    ws.append([
        'Segment', 'Campaign code', 'Phones in segment',
        'Bookings created in window',
        'Currently Confirmed (BK)', 'Currently On Offer (OF)',
        'Currently Cancelled (CX)',
        'Transferred OF -> BK',
        'Used campaign code',
    ])
    for _, label in SEGMENT_FILES:
        m = seg_metrics[label]
        ws.append([
            label,
            SEGMENT_CODES.get(label, ''),
            seg_counts[label],
            len(m['created']),
            len(m['cur_bk']),
            len(m['cur_of']),
            len(m['cur_cx']),
            len(m['transferred_to_bk']),
            len(m['with_campaign_code']),
        ])
    ws.append([])
    ws.append(['Match rule:', 'last 9 digits of PHONE_NUMBER'])
    style_header(ws, 9)
    autosize(ws)

    # Sheet 2 — Segment Details
    ws = wb.create_sheet('Segment Details')
    ws.append([
        'Segment', 'Expected campaign code', 'RES_ID', 'Current status',
        'Created in window', 'Transferred OF -> BK',
        'OF -> BK at',
        'Campaign codes applied', 'Used campaign code?',
        'PHONE', 'EMAIL', 'FULL_NAME', 'RES_INIT_DATE',
        'PACKAGE_TYPE', 'SAIL_DATE_FROM',
    ])
    campaign_only = set(ALL_CAMPAIGN_CODES)
    for _, label in SEGMENT_FILES:
        expected = SEGMENT_CODES.get(label, '')
        seg_res_ids = sorted({
            res_id for res_id, rec in res_info.items()
            if label in phone_to_segments.get(rec['phone'], set())
        })
        for res_id in seg_res_ids:
            rec = res_info[res_id]
            applied = promos.get(res_id, set())
            codes = sorted(applied & campaign_only)
            ws.append([
                label, expected, res_id,
                rec['cur_status'] or rec['latest_event_new_status'] or '',
                'Y' if rec['created_in_window'] else 'N',
                'Y' if rec['transferred_to_bk'] else 'N',
                rec['To_BK'].strftime('%Y-%m-%d %H:%M:%S') if rec['To_BK'] else '',
                '|'.join(codes),
                'Y' if expected and expected in codes else 'N',
                rec['phone'] or '',
                rec['email'] or '',
                rec['name'] or '',
                rec['init_date'].strftime('%Y-%m-%d %H:%M:%S') if rec['init_date'] else '',
                rec['pkg'] or '',
                rec['sail'].strftime('%Y-%m-%d') if rec['sail'] else '',
            ])
    style_header(ws, 15)
    autosize(ws)

    # Sheet 3 — Code Bookings (every booking with a campaign code)
    ws = wb.create_sheet('Code Bookings')
    ws.append([
        'Campaign code', 'Mapped segments', 'RES_ID', 'Current status',
        'PHONE', 'EMAIL', 'Phone in mapped segment?', 'Phone in other segment?',
        'FULL_NAME', 'RES_INIT_DATE', 'PACKAGE_TYPE', 'SAIL_DATE_FROM',
    ])
    seen = set()
    for row in sorted(code_bookings, key=lambda r: (r['code'], r['res_id'])):
        res_id = row['res_id']
        if res_id in seen:
            continue
        seen.add(res_id)
        code = row['code'].upper()
        mapped_segs = CODE_SEGMENTS.get(code, [])
        phone = row['phone'] or ''
        phone_segs = phone_to_segments.get(phone, set())
        in_mapped = bool(phone_segs & set(mapped_segs))
        in_other = bool(phone_segs - set(mapped_segs))
        ws.append([
            code,
            ' / '.join(mapped_segs),
            res_id,
            row['status'] or '',
            phone,
            row['email'] or '',
            'Y' if in_mapped else 'N',
            'Y' if in_other else 'N',
            row['name'] or '',
            row['init_date'].strftime('%Y-%m-%d %H:%M:%S') if row['init_date'] else '',
            row['pkg'] or '',
            row['sail'].strftime('%Y-%m-%d') if row['sail'] else '',
        ])
    ws.append([])
    ws.append(['─── Totals by code ───'])
    ws.append(['Code', 'Mapped segments', 'Bookings tagged',
               'Matched a mapped-segment phone', 'Matched another-segment phone',
               'Unknown phone (not in any segment)'])
    for code in ALL_CAMPAIGN_CODES:
        mapped_segs = CODE_SEGMENTS[code]
        res_ids_for_code = {r['res_id'] for r in code_bookings if r['code'].upper() == code}
        in_mapped = 0
        in_other = 0
        unknown = 0
        for res_id in res_ids_for_code:
            phones_on_booking = {r['phone'] for r in code_bookings
                                 if r['code'].upper() == code and r['res_id'] == res_id and r['phone']}
            segs_for_booking = set()
            for p in phones_on_booking:
                segs_for_booking |= phone_to_segments.get(p, set())
            if segs_for_booking & set(mapped_segs):
                in_mapped += 1
            elif segs_for_booking:
                in_other += 1
            else:
                unknown += 1
        ws.append([
            code,
            ' / '.join(mapped_segs),
            len(res_ids_for_code),
            in_mapped,
            in_other,
            unknown,
        ])
    style_header(ws, 12)
    autosize(ws)

    today = datetime.date.today().isoformat()
    xlsx_path = os.path.join(OUT_DIR, f'May26 Repeaters Campaign {today}.xlsx')
    wb.save(xlsx_path)
    print(f'Wrote {xlsx_path}')


if __name__ == '__main__':
    main()
